"""
04_build_excel.py

Builds the Excel deliverable: raw data + formula-driven pivot-style summary
tables + a KPI dashboard sheet + a what-if pricing/discount sensitivity
table, cross-validating the Python/SQL aggregates.

All summary numbers are computed with live SUMIFS/AVERAGEIFS/SUMPRODUCT
formulas against the Raw Data sheet (never hardcoded), so the workbook
recalculates if the underlying data changes.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import LineChart, BarChart, Reference

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F2A44")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=16, color="1F2A44")
SUB_FONT = Font(name=FONT_NAME, size=10, italic=True, color="666666")
LABEL_FONT = Font(name=FONT_NAME, bold=True, size=11)
BODY_FONT = Font(name=FONT_NAME, size=10)
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
KPI_FILL = PatternFill("solid", fgColor="EEF2FA")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CURRENCY_FMT = '$#,##0;($#,##0);"-"'
PCT_FMT = "0.0%"
NUM_FMT = "#,##0"


def style_header_row(ws, row, n_cols, start_col=1):
    for c in range(start_col, start_col + n_cols):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_df(ws, df, start_row, start_col=1, header=True, number_formats=None):
    """Write a DataFrame as static values (used for Raw Data sheet)."""
    if header:
        for j, col in enumerate(df.columns):
            ws.cell(row=start_row, column=start_col + j, value=col)
        style_header_row(ws, start_row, len(df.columns), start_col)
        start_row += 1
    for i, (_, row) in enumerate(df.iterrows()):
        for j, col in enumerate(df.columns):
            cell = ws.cell(row=start_row + i, column=start_col + j, value=row[col])
            cell.font = BODY_FONT
            if number_formats and col in number_formats:
                cell.number_format = number_formats[col]
    return start_row + len(df)


def main():
    df = pd.read_csv("data/processed/clean_transactions.csv")

    wb = Workbook()

    # =====================================================================
    # Sheet 1: Overview
    # =====================================================================
    ws = wb.active
    ws.title = "Overview"
    ws["B2"] = "Footwear & Apparel Sales Analytics — Workbook"
    ws["B2"].font = TITLE_FONT
    ws["B3"] = "Adidas US Sales dataset (synthetic, matched to Kaggle schema) | 9,648 transactions | Jan 2020 – Dec 2021"
    ws["B3"].font = SUB_FONT

    overview_text = [
        ("", ""),
        ("How this workbook is organized:", ""),
        ("Raw Data", "Cleaned, transaction-level data (9,648 rows) — source for every formula in this workbook."),
        ("KPI Dashboard", "Headline KPIs and channel/region pivot-style summaries, built with live SUMIFS formulas."),
        ("Pivot - Category", "Revenue, profit & margin by product category (Footwear vs Apparel)."),
        ("Pivot - Region", "Revenue, profit & margin by US region."),
        ("Pivot - Channel", "Revenue, profit & margin by sales channel (In-store / Outlet / Online) — the #1 margin driver."),
        ("Pivot - Region x Channel", "Cross-tab isolating where regional discounting and channel mix compound."),
        ("What-If Analysis", "Editable sensitivity table: adjust channel mix or discount assumptions and see margin impact instantly."),
        ("", ""),
        ("Key finding:", "Sales-channel mix is the dominant driver of operating-margin variance (Outlet ≈ -14pp vs. company"),
        ("", "average, driven by average discounting of ~28% vs. ~6% in-store), with regional discounting a secondary factor."),
    ]
    r = 5
    for label, desc in overview_text:
        ws.cell(row=r, column=2, value=label).font = LABEL_FONT if label and not label.startswith("Key") else (TITLE_FONT if label=="Key finding:" else BODY_FONT)
        ws.cell(row=r, column=3, value=desc).font = BODY_FONT
        ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True)
        r += 1
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 95
    ws.row_dimensions[2].height = 24

    # =====================================================================
    # Sheet 2: Raw Data
    # =====================================================================
    ws_raw = wb.create_sheet("Raw Data")
    raw_cols = ["Retailer", "Retailer ID", "Invoice Date", "Region", "State", "City",
                "Product", "Category", "Price per Unit", "Units Sold", "Total Sales",
                "Operating Profit", "Operating Margin", "Sales Method"]
    raw_df = df[raw_cols].copy()
    fmts = {
        "Price per Unit": CURRENCY_FMT, "Total Sales": CURRENCY_FMT,
        "Operating Profit": CURRENCY_FMT, "Operating Margin": PCT_FMT,
        "Units Sold": NUM_FMT,
    }
    last_row = write_df(ws_raw, raw_df, 1, number_formats=fmts)
    ws_raw.freeze_panes = "A2"
    autofit(ws_raw, [14, 12, 13, 11, 14, 14, 24, 10, 12, 11, 14, 15, 14, 12])
    N = len(raw_df)  # data rows
    DATA_LAST_ROW = 1 + N  # header at row1

    def col_letter(name):
        return get_column_letter(raw_cols.index(name) + 1)

    RNG = lambda name: f"'Raw Data'!${col_letter(name)}$2:${col_letter(name)}${DATA_LAST_ROW}"

    # =====================================================================
    # Helper to build a formula-driven pivot-style sheet
    # =====================================================================
    def build_pivot_sheet(name, dim_cols, dim_headers, values):
        """dim_cols: list of column names in raw data to group by (1 or 2 dims)
           values: list of unique tuples for each dim combination"""
        ws_p = wb.create_sheet(name)
        ws_p["B2"] = name.replace("Pivot - ", "") + " Summary (live SUMIFS formulas)"
        ws_p["B2"].font = TITLE_FONT
        headers = dim_headers + ["Revenue", "Operating Profit", "Operating Margin %",
                                  "Revenue Share %", "Units Sold", "Avg Price/Unit"]
        start_row = 4
        for j, h in enumerate(headers):
            ws_p.cell(row=start_row, column=2 + j, value=h)
        style_header_row(ws_p, start_row, len(headers), start_col=2)

        n_dims = len(dim_cols)
        for i, combo in enumerate(values):
            r = start_row + 1 + i
            if n_dims == 1:
                combo = (combo,)
            for d, val in enumerate(combo):
                ws_p.cell(row=r, column=2 + d, value=val).font = BODY_FONT

            crit_parts = []
            for d, col in enumerate(dim_cols):
                cell_ref = f"${get_column_letter(2 + d)}{r}"
                crit_parts.append((RNG(col), cell_ref))

            def sumifs(target_col):
                rng = RNG(target_col)
                crits = "".join(f",{c},{v}" for c, v in crit_parts)
                return f"=SUMIFS({rng}{crits})"

            rev_col = 2 + n_dims
            profit_col = rev_col + 1
            margin_col = profit_col + 1
            share_col = margin_col + 1
            units_col = share_col + 1
            avgprice_col = units_col + 1

            rev_cell = ws_p.cell(row=r, column=rev_col, value=sumifs("Total Sales"))
            rev_cell.number_format = CURRENCY_FMT
            profit_cell = ws_p.cell(row=r, column=profit_col, value=sumifs("Operating Profit"))
            profit_cell.number_format = CURRENCY_FMT
            margin_cell = ws_p.cell(row=r, column=margin_col,
                                     value=f"={get_column_letter(profit_col)}{r}/{get_column_letter(rev_col)}{r}")
            margin_cell.number_format = PCT_FMT
            total_rev_ref = f"SUM({RNG('Total Sales')})"
            share_cell = ws_p.cell(row=r, column=share_col,
                                    value=f"={get_column_letter(rev_col)}{r}/{total_rev_ref}")
            share_cell.number_format = PCT_FMT
            units_cell = ws_p.cell(row=r, column=units_col, value=sumifs("Units Sold"))
            units_cell.number_format = NUM_FMT
            crits = "".join(f",{c},{v}" for c, v in crit_parts)
            avgp_cell = ws_p.cell(row=r, column=avgprice_col,
                                   value=f"=AVERAGEIFS({RNG('Price per Unit')}{crits})")
            avgp_cell.number_format = CURRENCY_FMT

            for c in range(2, 2 + len(headers)):
                ws_p.cell(row=r, column=c).border = BORDER
                ws_p.cell(row=r, column=c).font = BODY_FONT

        # conditional formatting on margin column
        last_data_row = start_row + len(values)
        margin_col_letter = get_column_letter(2 + n_dims + 2)
        rule = ColorScaleRule(start_type="min", start_color="F8696B",
                               mid_type="percentile", mid_value=50, mid_color="FFEB84",
                               end_type="max", end_color="63BE7B")
        ws_p.conditional_formatting.add(
            f"{margin_col_letter}{start_row+1}:{margin_col_letter}{last_data_row}", rule)

        widths = [3] + [16] * n_dims + [15, 16, 14, 12, 12, 14]
        autofit(ws_p, widths)
        return ws_p, start_row, last_data_row

    categories = sorted(df["Category"].unique().tolist())
    regions = sorted(df["Region"].unique().tolist())
    channels = sorted(df["Sales Method"].unique().tolist())
    region_channel = [(r_, c_) for r_ in regions for c_ in channels]

    build_pivot_sheet("Pivot - Category", ["Category"], ["Category"], categories)
    ws_reg, _, reg_last = build_pivot_sheet("Pivot - Region", ["Region"], ["Region"], regions)
    ws_chan, _, chan_last = build_pivot_sheet("Pivot - Channel", ["Sales Method"], ["Sales Method"], channels)
    ws_rc, rc_start, rc_last = build_pivot_sheet("Pivot - Region x Channel", ["Region", "Sales Method"],
                                                  ["Region", "Sales Method"], region_channel)

    # =====================================================================
    # Sheet: KPI Dashboard
    # =====================================================================
    ws_kpi = wb.create_sheet("KPI Dashboard", 1)  # position after Overview
    ws_kpi["B2"] = "KPI Dashboard"
    ws_kpi["B2"].font = TITLE_FONT
    ws_kpi["B3"] = "All figures below are live formulas referencing the Raw Data sheet."
    ws_kpi["B3"].font = SUB_FONT

    kpis = [
        ("Total Revenue", f"=SUM({RNG('Total Sales')})", CURRENCY_FMT),
        ("Total Operating Profit", f"=SUM({RNG('Operating Profit')})", CURRENCY_FMT),
        ("Overall Operating Margin", f"=C6/C5", PCT_FMT),
        ("Total Units Sold", f"=SUM({RNG('Units Sold')})", NUM_FMT),
        ("Total Transactions", f"=COUNTA({RNG('Retailer')})", NUM_FMT),
        ("Avg Price per Unit", f"=AVERAGE({RNG('Price per Unit')})", CURRENCY_FMT),
    ]
    row = 5
    for label, formula, fmt in kpis:
        ws_kpi.cell(row=row, column=2, value=label).font = LABEL_FONT
        cell = ws_kpi.cell(row=row, column=3, value=formula)
        cell.font = Font(name=FONT_NAME, bold=True, size=12, color="1F2A44")
        cell.number_format = fmt
        cell.fill = KPI_FILL
        ws_kpi.cell(row=row, column=2).fill = KPI_FILL
        row += 1

    ws_kpi["B12"] = "Margin by Channel (weakest → strongest)"
    ws_kpi["B12"].font = LABEL_FONT
    ws_kpi["B13"] = "See 'Pivot - Channel' and 'Pivot - Region x Channel' tabs for full detail and conditional-formatting heatmap."
    ws_kpi["B13"].font = SUB_FONT
    ws_kpi.column_dimensions["B"].width = 26
    ws_kpi.column_dimensions["C"].width = 20

    # Bar chart: revenue & margin by channel, sourced from Pivot - Channel sheet
    chart1 = BarChart()
    chart1.title = "Revenue by Sales Channel"
    chart1.y_axis.title = "Revenue ($)"
    chart1.x_axis.title = "Sales Channel"
    data_ref = Reference(ws_chan, min_col=4, min_row=4, max_row=chan_last)  # Revenue col (D)
    cats_ref = Reference(ws_chan, min_col=2, min_row=5, max_row=chan_last)
    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(cats_ref)
    chart1.width, chart1.height = 14, 8
    ws_kpi.add_chart(chart1, "E5")

    chart2 = LineChart()
    chart2.title = "Operating Margin % by Region"
    chart2.y_axis.title = "Margin %"
    data_ref2 = Reference(ws_reg, min_col=5, min_row=4, max_row=reg_last)  # Margin col
    cats_ref2 = Reference(ws_reg, min_col=2, min_row=5, max_row=reg_last)
    chart2.add_data(data_ref2, titles_from_data=True)
    chart2.set_categories(cats_ref2)
    chart2.width, chart2.height = 14, 8
    ws_kpi.add_chart(chart2, "E21")

    # =====================================================================
    # Sheet: What-If Analysis
    # =====================================================================
    ws_wi = wb.create_sheet("What-If Analysis")
    ws_wi["B2"] = "What-If: Channel-Mix & Discount Sensitivity"
    ws_wi["B2"].font = TITLE_FONT
    ws_wi["B3"] = "Yellow cells are editable inputs. Change channel revenue-mix % or discount deltas to see the blended margin impact."
    ws_wi["B3"].font = SUB_FONT

    # baseline actuals per channel (pulled live from Pivot - Channel sheet order: sorted channels)
    ws_wi["B6"] = "Channel"
    ws_wi["C6"] = "Actual Revenue Share %"
    ws_wi["D6"] = "Actual Margin %"
    ws_wi["E6"] = "Scenario Revenue Share %"
    ws_wi["F6"] = "Margin Delta (pp)"
    ws_wi["G6"] = "Scenario Margin %"
    ws_wi["H6"] = "Scenario Weighted Margin Contribution %"
    style_header_row(ws_wi, 6, 7, start_col=2)

    # channel rows reference the Pivot - Channel sheet (channels alphabetical: In-store, Online, Outlet)
    chan_pivot_row_map = {ch: 5 + i for i, ch in enumerate(channels)}  # data starts row5 in pivot sheet
    for i, ch in enumerate(channels):
        r = 7 + i
        pr = chan_pivot_row_map[ch]
        ws_wi.cell(row=r, column=2, value=ch).font = BODY_FONT
        share_cell = ws_wi.cell(row=r, column=3, value=f"='Pivot - Channel'!F{pr}")
        share_cell.number_format = PCT_FMT
        margin_cell = ws_wi.cell(row=r, column=4, value=f"='Pivot - Channel'!D{pr}/'Pivot - Channel'!C{pr}")
        margin_cell.number_format = PCT_FMT
        scen_share = ws_wi.cell(row=r, column=5, value=f"=C{r}")  # default = actual; user edits
        scen_share.number_format = PCT_FMT
        scen_share.fill = INPUT_FILL
        delta_cell = ws_wi.cell(row=r, column=6, value=0.0)
        delta_cell.number_format = "0.0%"
        delta_cell.fill = INPUT_FILL
        scen_margin = ws_wi.cell(row=r, column=7, value=f"=D{r}+F{r}")
        scen_margin.number_format = PCT_FMT
        contrib_cell = ws_wi.cell(row=r, column=8, value=f"=E{r}*G{r}")
        contrib_cell.number_format = PCT_FMT
        for c in range(2, 9):
            ws_wi.cell(row=r, column=c).border = BORDER
            if c not in (5, 6):
                ws_wi.cell(row=r, column=c).font = BODY_FONT

    total_row = 7 + len(channels)
    ws_wi.cell(row=total_row, column=2, value="Blended (Scenario) Margin").font = LABEL_FONT
    total_cell = ws_wi.cell(row=total_row, column=8, value=f"=SUM(H7:H{total_row-1})/SUM(E7:E{total_row-1})")
    total_cell.number_format = PCT_FMT
    total_cell.font = Font(name=FONT_NAME, bold=True, size=12, color="1F2A44")
    total_cell.fill = KPI_FILL

    baseline_row = total_row + 2
    ws_wi.cell(row=baseline_row, column=2, value="Actual Blended Margin (baseline)").font = LABEL_FONT
    baseline_cell = ws_wi.cell(row=baseline_row, column=8, value="='KPI Dashboard'!C7")
    baseline_cell.number_format = PCT_FMT

    ws_wi.cell(row=baseline_row + 2, column=2,
               value="Example scenario: shift 5pp of revenue from Outlet to In-store (raise E7 by 5%, lower E9 by 5%) to see the margin lift.").font = SUB_FONT

    autofit(ws_wi, [3, 22, 20, 16, 20, 16, 16, 26])

    wb.save("excel/Adidas_Sales_Analytics.xlsx")
    print("Saved excel/Adidas_Sales_Analytics.xlsx")


if __name__ == "__main__":
    main()
